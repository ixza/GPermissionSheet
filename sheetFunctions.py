from google.oauth2 import service_account
from googleapiclient import discovery
from openpyxl.utils.cell import get_column_letter
import config
import os
import pygsheets
from string import digits


credentials = service_account.Credentials.from_service_account_file(config.CREDENTIALS_PATH)
service = discovery.build('sheets', 'v4', credentials=credentials)
gc = pygsheets.authorize(service_account_file=config.CREDENTIALS_PATH)
remove_digits = str.maketrans('', '', digits)

def add_element(dict, key, value):
   if key not in dict:
      dict[key] = []
   dict[key].append(value)
   
def set_element(dict, key, value):
   dict[key] = value
    
class sheetFunctions:
   protectedSheets = {} #title:range
   protectedSheetsId = {} #title:id
   users = {} #user:ranges
   sheetRanges = {} #sheetId:range
   sheetNumbers = 0
   
   def fetchData():

      request = service.spreadsheets().get(spreadsheetId=config.SPREADSHEET_ID, includeGridData=False)
      response = request.execute()
      sheetFunctions.sheetNumbers = len(response['sheets'])
      
      for i in range(0,sheetFunctions.sheetNumbers):
         if 'protectedRanges' in response['sheets'][i]:
            for y in range(0, len(response['sheets'][i]['protectedRanges'])):
               add_element(sheetFunctions.protectedSheets, response['sheets'][i]['properties']['title'], response['sheets'][i]['protectedRanges'][y]['range'])
               set_element(sheetFunctions.protectedSheetsId, response['sheets'][i]['properties']['title'], response['sheets'][i]['properties']['sheetId'])
               
               for x in range(0, len(response['sheets'][i]['protectedRanges'][y]['editors']['users'])):
                  add_element(sheetFunctions.users, response['sheets'][i]['protectedRanges'][y]['editors']['users'][x], response['sheets'][i]['protectedRanges'][y]['range'])
      
      return response
   
   #CONVERT R1C1 TO A1 NOTATION
   def a1conversion(username, sheetId):
      userListRange = []
      A = ""
      B = ""
      for i in range( 0, len(sheetFunctions.users[username])): 
            if sheetFunctions.users[username][i]['sheetId'] == sheetId:
               
               if 'startColumnIndex' in sheetFunctions.users[username][i]:
                  A = get_column_letter(sheetFunctions.users[username][i]['startColumnIndex']+1)
                  if 'startRowIndex' in sheetFunctions.users[username][i]:
                     A+= str(sheetFunctions.users[username][i]['startRowIndex']+1)
               
               if 'endColumnIndex' in sheetFunctions.users[username][i]:
                  B = get_column_letter(sheetFunctions.users[username][i]['endColumnIndex'])
                  if 'endRowIndex' in sheetFunctions.users[username][i]:
                     B+= str(sheetFunctions.users[username][i]['endRowIndex'])
               
               
               if A == "" or B =="":
                  userListRange.append('ALL')
               else:
                  userListRange.append(str(A) +':'+ str(B))
               userListRange.sort()
               
      return userListRange
   
   #FIX C1 CONVERSION
   def c1conversion(username, sheetId, rangeSet):
      userListRange = []
      A = ""
      B = ""
      for i in range( 0, len(sheetFunctions.users[username])): 
            if sheetFunctions.users[username][i]['sheetId'] == sheetId:
               
               if 'startColumnIndex' in sheetFunctions.users[username][i]:
                  A = sheetFunctions.users[username][i]['startColumnIndex']+1
               
               if 'endColumnIndex' in sheetFunctions.users[username][i]:
                  B = sheetFunctions.users[username][i]['endColumnIndex']
               
               if A == "" or B =="":
                  userListRange.append('ALL')
               else:
                  userListRange.append(str(A) +':'+ str(B))
                  rangeSet.add(get_column_letter(A) +'1:'+ get_column_letter(B)+'1')
               userListRange.sort()
      return userListRange
   
   #ADD BLANK INTO LIST TO PREVENT OUT OF RANGE IN SHEETS ERROR
   def addBlank(list):
    x = 0
    for i in range(0, len(list)):
        if len(list[i]) > x:
            x = i
    if len(list[0]) < x:
        for y in range(0, x):
            list[0].append(" ")
            
   #CONVERT EMAIL TO NAME
   def nameCheck(list, dict):
      namesList = []
      for i in range(0, len(list)):
         nameInList = False
         for key in dict:
            if list[i] in dict[key]:
               namesList.append(key)
               nameInList = True
               break
         if nameInList == False:
            namesList.append(" ")
      return namesList
            
   
   #OBTAIN ALL RANGES OF A USER
   def getUserRanges(sheetsIdList, userList):
      userListRanges = []
      for x in range (0, len(sheetsIdList)):
          userListRanges.append([])
          for i in range(0, len(userList)):
              ranges = sheetFunctions.a1conversion(userList[i],sheetsIdList[x])
              if ranges == []:
                  ranges = [' ']
              userListRanges[x].append(ranges)
              sheetFunctions.addBlank(userListRanges[x])
      return userListRanges